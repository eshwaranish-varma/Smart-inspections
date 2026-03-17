from __future__ import annotations
import logging
from io import BytesIO
from pathlib import Path
from typing import Optional

import openpyxl

logger = logging.getLogger(__name__)


class ExcelService:
    def __init__(self, file_path: Path):
        self.file_path = file_path
        self.column_map: dict[str, int] = {}
        self._detect_columns()

    def _detect_columns(self):
        """Read header row to map column names to indices."""
        if not self.file_path.exists():
            logger.warning("Excel file not found: %s", self.file_path)
            return

        wb = openpyxl.load_workbook(self.file_path, read_only=True)
        ws = wb.active
        headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.column_map = {h: i for i, h in enumerate(headers) if h}
        wb.close()
        logger.info("Excel columns detected: %s", list(self.column_map.keys()))

    def load_observations(self) -> list[dict]:
        """Read all rows from the Excel file."""
        if not self.file_path.exists():
            return []

        wb = openpyxl.load_workbook(self.file_path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        headers = list(self.column_map.keys())
        results = []
        for i, row in enumerate(rows):
            record: dict = {"id": i + 1}
            for j, val in enumerate(row):
                if j < len(headers):
                    key = headers[j].lower().replace(" ", "_").replace("/", "_")
                    record[key] = str(val) if val is not None else ""
            results.append(record)

        return results

    def get_statistics(self) -> dict:
        """Return summary statistics over the observation data."""
        obs = self.load_observations()
        if not obs:
            return {"total": 0}

        type_counts: dict[str, int] = {}
        cfr_counts: dict[str, int] = {}
        for o in obs:
            etype = o.get("establishment_type", o.get("type", "Unknown"))
            type_counts[etype] = type_counts.get(etype, 0) + 1
            cfr = o.get("cfr_citation", o.get("act_cfr_number", ""))
            if cfr:
                cfr_counts[cfr] = cfr_counts.get(cfr, 0) + 1

        top_cfr = sorted(cfr_counts.items(), key=lambda x: x[1], reverse=True)[:10]

        return {
            "total": len(obs),
            "by_establishment_type": type_counts,
            "by_cfr_part": cfr_counts,
            "top_citations": [{"citation": c, "count": n} for c, n in top_cfr],
            "column_names": list(self.column_map.keys()),
        }

    def search_observations(self, query: str = "", filters: dict | None = None) -> list[dict]:
        """Full-text search with optional field-level filters."""
        obs = self.load_observations()
        query_lower = query.lower()

        results = []
        for o in obs:
            if query_lower:
                text = " ".join(str(v) for v in o.values()).lower()
                if query_lower not in text:
                    continue

            if filters:
                skip = False
                for key, value in filters.items():
                    if value and str(o.get(key, "")).lower() != str(value).lower():
                        skip = True
                        break
                if skip:
                    continue

            results.append(o)

        return results

    def add_observation(self, data: dict) -> bool:
        """Append a new row to the Excel file."""
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active
            headers = list(self.column_map.keys())
            row = [data.get(h.lower().replace(" ", "_").replace("/", "_"), "") for h in headers]
            ws.append(row)
            wb.save(self.file_path)
            wb.close()
            return True
        except Exception as e:
            logger.error("Failed to add observation: %s", e)
            return False

    def export_filtered(self, observations: list[dict]) -> bytes:
        """Export observations list as an in-memory Excel file."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Filtered Observations"

        if observations:
            headers = list(observations[0].keys())
            ws.append(headers)
            for obs in observations:
                ws.append([obs.get(h, "") for h in headers])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
